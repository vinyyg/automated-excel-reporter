import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from pathlib import Path
from datetime import datetime
import config

VERDE    = PatternFill("solid", start_color="C6EFCE")
VERMELHO = PatternFill("solid", start_color="FFC7CE")
AMARELO  = PatternFill("solid", start_color="FFEB9C")

def abrir_excel(caminho):
    return load_workbook(caminho)

def ler_dados(caminho):
    df = pd.read_excel(caminho, header=1)
    df = df[:-1]
    df = df.dropna(subset=[df.columns[0]])
    return df

def aplicar_formula(df):
    df["Total (R$)"] = df["Quantidade"] * df["Valor Unit. (R$)"]
    return df

def aplicar_formatacao(ws):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status = row[6].value
        if status == "Aprovado":
            fill = VERDE
        elif status == "Cancelado":
            fill = VERMELHO
        elif status == "Pendente":
            fill = AMARELO
        else:
            continue
        for cell in row:
            cell.fill = fill

def salvar_excel(df, caminho):
    df.to_excel(caminho, index=False)

def obter_caminho_consolidado():
    data = datetime.now().strftime("%d-%m-%Y")
    pasta = Path(config.ARQUIVO_SAIDA).parent
    return pasta / f"Relatorio Consolidado - {data}.xlsx"

def copiar_para_consolidado(df, caminho_consolidado):
    if caminho_consolidado.exists():
        wb = load_workbook(caminho_consolidado)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(list(df.columns))

    for _, row in df.iterrows():
        ws.append(list(row))

    aplicar_formatacao(ws)
    wb.save(caminho_consolidado)

def processar(caminho_arquivo):
    df = ler_dados(caminho_arquivo)
    df = aplicar_formula(df)
    caminho_consolidado = obter_caminho_consolidado()
    copiar_para_consolidado(df, caminho_consolidado)
    return caminho_consolidado, df